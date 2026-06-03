#!/usr/bin/env python3
"""
Import all 177 portfolio projects from the Approach Media case-study Excel file
into PostgreSQL.  Also seeds the app_config table with the CDN base URL.

Usage:
    pip install openpyxl psycopg2-binary python-dotenv
    python scripts/import_from_excel.py --excel scripts/new_casestudy_step2_seo_ready.xlsx
    python scripts/import_from_excel.py --excel /path/to/file.xlsx --update   # overwrite existing slugs
    python scripts/import_from_excel.py --excel /path/to/file.xlsx --dry-run  # preview only

The script:
  1. Creates the app_config table row: key='media_cdn_base_url'
  2. For each Excel row: upserts Client, Exhibition, Industry(ies), StallType(s)
  3. Inserts Project + SeoMetadata + Media rows
  4. Stores image paths as RELATIVE paths (no domain) — the CDN base URL from
     app_config is prepended at render time, so switching CDN = one DB update.

Environment:
    DATABASE_URL  PostgreSQL connection string (read from .env by default)
"""

import argparse
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Missing: pip install openpyxl")

try:
    import psycopg2, psycopg2.extras
except ImportError:
    sys.exit("Missing: pip install psycopg2-binary")

try:
    from dotenv import load_dotenv
except ImportError:
    sys.exit("Missing: pip install python-dotenv")


# ── Config ─────────────────────────────────────────────────────

CDN_BASE_URL   = "https://pub-3142dbc1bfbb47b191e0dca72e867a0f.r2.dev"
CDN_DESCRIPTION = (
    "Base URL for all media assets on Cloudflare R2. "
    "Update this one value to migrate every project image to a new CDN — "
    "no project records need to change."
)

EXCEL_COLS = [
    "title", "slug", "client_name", "exhibition_name", "venue_name",
    "city", "description", "country", "stall_area_sqm", "stall_area_sqft",
    "stall_height_m", "floors", "build_year", "design_brief", "ai_summary",
    "design_style", "materials_used", "special_features", "awards",
    "industries", "stall_types", "status", "is_featured", "display_order",
    "meta_title", "meta_description", "og_title", "og_description",
    "hero_image_url", "hero_image_alt", "hero_image_caption", "gallery_images",
]


# ── Helpers ────────────────────────────────────────────────────

def slugify(text: str) -> str:
    t = text.lower().strip()
    t = re.sub(r"[^\w\s-]", "", t)
    t = re.sub(r"[\s_-]+", "-", t)
    return t.strip("-")[:220]


def to_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def to_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("1", "true", "yes", "y")


def split_pipe(v) -> list[str]:
    if not v:
        return []
    return [x.strip() for x in re.split(r"[|,]", str(v)) if x.strip()]


def clean_path(p: str) -> str:
    """Strip any accidental full CDN URL — store only the relative path."""
    p = p.strip()
    for prefix in (CDN_BASE_URL, "https://", "http://"):
        if p.startswith(prefix):
            # remove the protocol+domain part
            p = re.sub(r"^https?://[^/]+/?", "", p)
    return p.lstrip("/")


def parse_gallery(cell) -> list[dict]:
    if not cell:
        return []
    items = []
    for part in re.split(r",\s*(?=\d{4}/)", str(cell)):
        path = clean_path(part)
        if path:
            items.append({"path": path, "alt": "", "caption": None})
    return items


# ── DB helpers ─────────────────────────────────────────────────

def get_or_create(cur, table: str, name_col: str, name: str, extra: dict = {}) -> int:
    cur.execute(f"SELECT id FROM {table} WHERE {name_col} = %s", (name,))
    row = cur.fetchone()
    if row:
        return row[0]
    slug = slugify(name)
    # ensure unique slug
    base = slug
    i = 2
    while True:
        cur.execute(f"SELECT id FROM {table} WHERE slug = %s", (slug,))
        if not cur.fetchone():
            break
        slug = f"{base}-{i}"
        i += 1
    cols = f"{name_col}, slug" + (", " + ", ".join(extra.keys()) if extra else "")
    placeholders = "%s, %s" + (", " + ", ".join(["%s"] * len(extra)) if extra else "")
    vals = (name, slug) + tuple(extra.values())
    cur.execute(f"INSERT INTO {table} ({cols}) VALUES ({placeholders}) RETURNING id", vals)
    return cur.fetchone()[0]


# ── Main import ────────────────────────────────────────────────

def seed_cdn_config(cur):
    cur.execute("""
        INSERT INTO app_config (key, value, description, updated_at)
        VALUES (%s, %s, %s, now())
        ON CONFLICT (key) DO UPDATE
          SET value       = EXCLUDED.value,
              description = EXCLUDED.description,
              updated_at  = now()
    """, ("media_cdn_base_url", CDN_BASE_URL, CDN_DESCRIPTION))
    print(f"  CDN config set → {CDN_BASE_URL}")


def import_row(cur, data: dict, update: bool) -> str:
    slug = (data.get("slug") or slugify(data["title"])).strip()

    cur.execute("SELECT id FROM projects WHERE slug = %s", (slug,))
    existing = cur.fetchone()
    if existing and not update:
        return f"SKIP  {slug}"

    client_id = None
    if data.get("client_name"):
        client_id = get_or_create(cur, "clients", "name", data["client_name"])

    exhibition_id = None
    if data.get("exhibition_name"):
        extra = {}
        if data.get("venue_name"):  extra["venue_name"] = data["venue_name"]
        if data.get("city"):        extra["city"]       = data["city"]
        if data.get("country"):     extra["country"]    = data["country"] or "India"
        exhibition_id = get_or_create(cur, "exhibitions", "name", data["exhibition_name"], extra)

    def json_list(v):
        items = split_pipe(v)
        return json.dumps(items) if items else None

    project_data = {
        "title":           data["title"],
        "slug":            slug,
        "stall_area_sqm":  to_float(data.get("stall_area_sqm")),
        "stall_area_sqft": to_float(data.get("stall_area_sqft")),
        "stall_height_m":  to_float(data.get("stall_height_m")),
        "floors":          to_int(data.get("floors")) or 1,
        "build_year":      to_int(data.get("build_year")),
        "city":            data.get("city") or None,
        "exhibition_id":   exhibition_id,
        "client_id":       client_id,
        "description":     data.get("description") or "",
        "design_brief":    data.get("design_brief") or None,
        "ai_summary":      data.get("ai_summary") or None,
        "design_style":    data.get("design_style") or None,
        "materials_used":  json_list(data.get("materials_used")),
        "special_features":json_list(data.get("special_features")),
        "awards":          json_list(data.get("awards")),
        "status":          data.get("status") or "published",
        "is_featured":     to_bool(data.get("is_featured")),
        "display_order":   to_int(data.get("display_order")) or 0,
    }

    if existing:
        project_id = existing[0]
        set_parts = [f"{k} = %({k})s" for k in project_data if k != "slug"]
        cur.execute(
            f"UPDATE projects SET {', '.join(set_parts)}, updated_at = now() WHERE id = %(id)s",
            {**project_data, "id": project_id},
        )
        action = "UPDATE"
    else:
        cols   = ", ".join(project_data.keys())
        params = ", ".join(f"%({k})s" for k in project_data)
        cur.execute(f"INSERT INTO projects ({cols}) VALUES ({params}) RETURNING id", project_data)
        project_id = cur.fetchone()[0]
        action = "INSERT"

    # SEO metadata
    seo = {
        "project_id":       project_id,
        "meta_title":       data.get("meta_title") or None,
        "meta_description": data.get("meta_description") or None,
        "og_title":         data.get("og_title") or None,
        "og_description":   data.get("og_description") or None,
        "og_image_url":     clean_path(data.get("hero_image_url") or "") or None,
    }
    if any(v for k, v in seo.items() if k != "project_id"):
        cur.execute("""
            INSERT INTO seo_metadata
              (project_id, meta_title, meta_description, og_title, og_description, og_image_url)
            VALUES
              (%(project_id)s, %(meta_title)s, %(meta_description)s,
               %(og_title)s, %(og_description)s, %(og_image_url)s)
            ON CONFLICT (project_id) DO UPDATE SET
              meta_title       = EXCLUDED.meta_title,
              meta_description = EXCLUDED.meta_description,
              og_title         = EXCLUDED.og_title,
              og_description   = EXCLUDED.og_description,
              og_image_url     = EXCLUDED.og_image_url,
              updated_at       = now()
        """, seo)

    # Media — clear and re-insert on update
    if action == "UPDATE":
        cur.execute("DELETE FROM media WHERE project_id = %s", (project_id,))

    order = 0
    hero_path = clean_path(data.get("hero_image_url") or "")
    if hero_path:
        cur.execute("""
            INSERT INTO media
              (project_id, media_type, url, alt_text, caption, display_order, is_hero, is_thumbnail)
            VALUES (%s, 'image', %s, %s, %s, %s, true, true)
        """, (
            project_id,
            hero_path,
            (data.get("hero_image_alt") or data["title"] + " exhibition stall").strip(),
            data.get("hero_image_caption") or None,
            order,
        ))
        order += 1

    # Gallery — comma-separated paths
    gallery = parse_gallery(data.get("gallery_images"))
    # skip index 0 if it's the same path as hero (already inserted)
    for img in gallery:
        if img["path"] == hero_path and order == 1:
            continue  # hero already inserted as order=0
        media_type = "render_3d" if "render" in img["path"].lower() else "image"
        alt = data["title"] + " exhibition stall photo " + str(order)
        cur.execute("""
            INSERT INTO media (project_id, media_type, url, alt_text, caption, display_order)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (project_id, media_type, img["path"], alt, img["caption"], order))
        order += 1

    # Industries
    cur.execute("DELETE FROM project_industries WHERE project_id = %s", (project_id,))
    for i, name in enumerate(split_pipe(data.get("industries"))):
        ind_id = get_or_create(cur, "industries", "name", name)
        cur.execute("""
            INSERT INTO project_industries (project_id, industry_id, is_primary)
            VALUES (%s, %s, %s) ON CONFLICT DO NOTHING
        """, (project_id, ind_id, i == 0))

    # Stall types
    cur.execute("DELETE FROM project_stall_types WHERE project_id = %s", (project_id,))
    for i, name in enumerate(split_pipe(data.get("stall_types"))):
        st_id = get_or_create(cur, "stall_types", "name", name)
        cur.execute("""
            INSERT INTO project_stall_types (project_id, stall_type_id, is_primary)
            VALUES (%s, %s, %s) ON CONFLICT DO NOTHING
        """, (project_id, st_id, i == 0))

    return f"{action} {slug}"


def main():
    parser = argparse.ArgumentParser(description="Import portfolio projects from Excel into PostgreSQL")
    parser.add_argument("--excel",   required=True, help="Path to .xlsx file")
    parser.add_argument("--env",     default=".env", help=".env file (default: .env in cwd)")
    parser.add_argument("--update",  action="store_true", help="Overwrite existing slugs")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, no DB writes")
    args = parser.parse_args()

    load_dotenv(args.env)
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        sys.exit("DATABASE_URL not set in env / .env file")

    wb = openpyxl.load_workbook(args.excel)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        if d.get("title"):
            rows.append(d)

    print(f"Loaded {len(rows)} project(s) from {args.excel}")

    if args.dry_run:
        for r in rows:
            print(f"  DRY-RUN: '{r['title'][:60]}' → {r.get('slug','')}")
        return

    conn = psycopg2.connect(db_url)
    conn.autocommit = False
    cur = conn.cursor()

    # Seed CDN config
    seed_cdn_config(cur)
    conn.commit()

    ok = err = skip = 0
    for r in rows:
        try:
            result = import_row(cur, r, args.update)
            conn.commit()
            print(f"  {result}")
            if result.startswith("SKIP"):
                skip += 1
            else:
                ok += 1
        except Exception as e:
            conn.rollback()
            print(f"  ERROR '{str(r.get('title','?'))[:50]}': {e}")
            err += 1

    cur.close()
    conn.close()
    print(f"\nDone — {ok} imported, {skip} skipped, {err} errors")
    if ok:
        print(f"\nCDN base URL stored in DB: {CDN_BASE_URL}")
        print("To change CDN later, run:")
        print("  UPDATE app_config SET value = 'https://new-cdn.com' WHERE key = 'media_cdn_base_url';")


if __name__ == "__main__":
    main()
